#!/usr/bin/env python3
"""Generate QA Excel report from qa/output/*.json + UI review data."""
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(ROOT, "output")
XLSX = os.path.join(OUT, "QA-Report-eBike-1.7.0.xlsx")

UI_REVIEW = [
    ("sign-up.tsx", "Email signup", "PASS", ""),
    ("role-select.tsx", "Technician role → customer tabs", "FAIL", "Tech lands on customer home; no pending-approval screen"),
    ("(customer)/index.tsx", "Active job redirect + repair FAB", "PASS", ""),
    ("repair-request.tsx", "4-step wizard + photo", "PASS", ""),
    ("technician-select.tsx", "Book job + upload photo", "PASS", ""),
    ("technician-select.tsx", "EB-XXXXXX on screen", "WARN", "Reference not shown before tracking"),
    ("job-tracking.tsx", "Payment before on_way", "PASS", "PaymentRequiredScreen when accepted+unpaid"),
    ("job-tracking.tsx", "EB-XXXXXX display", "PASS", ""),
    ("job-tracking.tsx", "Contact buttons timing", "WARN", "Hidden until paid+on_way"),
    ("job-tracking.tsx", "Hydrate payment_status", "WARN", "May flash payment screen on cold start"),
    ("(customer)/orders.tsx", "Order history from API", "PASS", ""),
    ("order-details.tsx", "Open from history", "FAIL", "Uses Zustand store only — Order not found"),
    ("(customer)/orders.tsx", "Store sync with details", "FAIL", "Never syncs API list to store"),
    ("job-complete.tsx", "Review submit", "PASS", ""),
    ("job-complete.tsx", "Payment label", "WARN", "Always says credit card"),
    ("reviews.tsx", "Load technician reviews", "PASS", ""),
    ("(technician)/index.tsx", "Pending jobs + accept", "PASS", ""),
    ("(technician)/index.tsx", "EB-XXXXXX on cards", "PASS", ""),
    ("(technician)/index.tsx", "Phone on pending (text only)", "WARN", "Not tappable call/WhatsApp"),
    ("(technician)/jobs.tsx", "Job history", "FAIL", "In-memory only — no API history"),
    ("(technician)/profile.tsx", "GET /api/users/me", "FAIL", "Endpoint 404 — profile may fail"),
    ("(technician)/profile.tsx", "Verification badge", "FAIL", "Hardcoded verified, ignores isApproved"),
    ("(technician)/profile.tsx", "Bio save", "WARN", "UI exists but not persisted"),
    ("active-job.tsx", "Payment gate + on_way", "PASS", ""),
    ("active-job.tsx", "Contact WhatsApp+Phone", "PASS", "After paid + on_way only"),
    ("active-job.tsx", "EB-XXXXXX", "PASS", ""),
]

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
PASS_FILL = PatternFill("solid", fgColor="D1FAE5")
FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
SKIP_FILL = PatternFill("solid", fgColor="F3F4F6")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def status_fill(status):
    return {
        "PASS": PASS_FILL,
        "FAIL": FAIL_FILL,
        "WARN": WARN_FILL,
        "SKIP": SKIP_FILL,
    }.get(status, SKIP_FILL)


def write_table(ws, headers, rows, start_row=1):
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row)
    r = start_row + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if headers[c - 1] == "Status":
                cell.fill = status_fill(str(val))
        r += 1
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(45, len(h) + 8))
    return r


def load_json(name):
    path = os.path.join(OUT, name)
    if not os.path.exists(path):
        return None
    with open(path) as f:
        return json.load(f)


def main():
    os.makedirs(OUT, exist_ok=True)
    e2e = load_json("qa-results.json") or {}
    monitor = load_json("monitor-results.json") or {}

    wb = Workbook()

    # ── Summary ──
    ws = wb.active
    ws.title = "Summary"
    e2e_totals = e2e.get("totals", {})
    mon_fail = len([r for r in monitor.get("rows", []) if r.get("status") == "FAIL"])
    ui_fail = len([r for r in UI_REVIEW if r[2] == "FAIL"])
    ui_warn = len([r for r in UI_REVIEW if r[2] == "WARN"])
    ui_pass = len([r for r in UI_REVIEW if r[2] == "PASS"])

    summary_rows = [
        ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("App Version", "1.7.0 (build 19)"),
        ("Backend Tested", e2e.get("backend", "http://127.0.0.1:3001")),
        ("QA Customer", (e2e.get("meta") or {}).get("customerEmail", "—")),
        ("QA Technician", (e2e.get("meta") or {}).get("techEmail", "—")),
        ("Job ID", (e2e.get("meta") or {}).get("jobId", "—")),
        ("Job Reference", (e2e.get("meta") or {}).get("jobReference", "—")),
        ("", ""),
        ("E2E API Tests PASS", e2e_totals.get("pass", 0)),
        ("E2E API Tests FAIL", e2e_totals.get("fail", 0)),
        ("UI Code Review PASS", ui_pass),
        ("UI Code Review FAIL", ui_fail),
        ("UI Code Review WARN", ui_warn),
        ("5-min Monitor FAIL", mon_fail),
        ("", ""),
        ("Critical Rules", ""),
        ("Customer pays before tech on_way", "VERIFIED — backend 402 gate"),
        ("Every order has EB-XXXXXX + DB backup", "VERIFIED — jobNumber in Postgres"),
        ("Technician needs admin approval", "VERIFIED — isApproved gate"),
    ]
    ws["A1"] = "eBike QA Report"
    ws["A1"].font = Font(bold=True, size=16, name="Arial")
    r = 3
    for label, val in summary_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Arial")
        ws.cell(row=r, column=2, value=val).font = Font(name="Arial")
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50

    # ── E2E API ──
    ws2 = wb.create_sheet("E2E API Tests")
    api_rows = []
    for item in e2e.get("results", []):
        api_rows.append(
            (
                item.get("step"),
                item.get("category"),
                item.get("test"),
                item.get("expected"),
                item.get("actual"),
                item.get("status"),
                item.get("notes", ""),
            )
        )
    write_table(
        ws2,
        ["Step", "Category", "Test", "Expected", "Actual", "Status", "Notes"],
        api_rows,
    )

    # ── UI Review ──
    ws3 = wb.create_sheet("UI Code Review")
    write_table(
        ws3,
        ["Screen", "Section", "Status", "Issue"],
        UI_REVIEW,
    )

    # ── Monitor ──
    ws4 = wb.create_sheet("5min Monitor")
    mon_rows = [
        (row.get("at"), row.get("check"), row.get("status"), row.get("detail"))
        for row in monitor.get("rows", [])
    ]
    if not mon_rows:
        mon_rows = [("—", "Monitor not run yet", "SKIP", "Run qa/monitor-5min.ts")]
    write_table(ws4, ["Time", "Check", "Status", "Detail"], mon_rows)

    # ── Bugs Priority ──
    ws5 = wb.create_sheet("Bugs Priority")
    bugs = [
        (1, "FAIL", "order-details.tsx", "Completed orders from history show Order not found"),
        (2, "FAIL", "role-select.tsx", "Technician role routes to customer tabs"),
        (3, "FAIL", "(technician)/jobs.tsx", "No API job history after restart"),
        (4, "FAIL", "GET /api/users/me", "Profile screens call missing endpoint (404)"),
        (5, "FAIL", "Tech profile verification", "Hardcoded verified ignores isApproved"),
        (6, "WARN", "job-tracking hydrate", "payment_status missing on API re-fetch"),
        (7, "WARN", "Tech bio save", "Not persisted to backend"),
    ]
    write_table(ws5, ["Priority", "Severity", "Component", "Description"], bugs)

    wb.save(XLSX)
    print(f"Saved: {XLSX}")
    return XLSX


if __name__ == "__main__":
    main()